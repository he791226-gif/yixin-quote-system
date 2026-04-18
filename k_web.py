import streamlit as st
import pandas as pd
from datetime import datetime
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment

# --- 1. 頁面基本設定 ---
st.set_page_config(page_title="翌新空壓機報價系統", layout="wide")
st.markdown("""
    <style>
    header, .stAppHeader, #MainMenu, footer, [data-testid="stDecoration"] {
        display: none !important;
        visibility: hidden !important;
    }
    [data-testid="stSidebarCollapsedControl"] { display: none !important; }
    .price-text { color: #E84118; font-size: 24px; font-weight: bold; }
    </style>
    """, unsafe_allow_html=True)

# --- 2. 產品規格介紹 (名稱已對齊) ---
product_specs = {
    "10馬力永磁變頻高效能補助專案型空壓機": (
        "型號:HCV-10PM-A\n馬力:1馬至10馬<隨壓力調整馬力>\n噪音:68±5\n"
        "電壓:三相220V\nIE4永磁高效率馬達<馬達無軸承><無皮帶>\n"
        "5kg~8kg可選變頻壓力\nLCD液晶顯示預警/警告/錯誤跳機保護\n"
        "外型尺寸:745*680*910(mm)\n變頻器與電機一體化非外掛變頻空壓機\n"
        "啟動電流衝擊小,恆壓恆溫控制,故障率低"
    ),
    "20馬力永磁變頻高效能補助專案型空壓機": (
        "型號:HCV-20PM-AA\n馬力:2馬至20馬<隨壓力調整馬力>\n噪音:70±5\n"
        "電壓:三相220V\nIE4永磁高效率馬達<馬達無軸承><無皮帶>\n"
        "5kg~8kg可選變頻壓力\nLCD液晶顯示預警/警告/錯誤跳機保護\n"
        "外型尺寸:1060*825*1120(mm)\n變頻器與電機一體化非外掛變頻空壓機\n"
        "啟動電流衝擊小,恆壓恆溫控制,故障率低"
    ),
    "30馬力永磁變頻高效能補助專案型空壓機": (
        "型號:HCV-30PM-A\n馬力:3馬至30馬<隨壓力調整馬力>\n噪音:75±5\n"
        "電壓:三相220V\nIE4永磁高效率馬達<馬達無軸承><無皮帶>\n"
        "5kg~8kg可選變頻壓力\nLCD液晶顯示預警/警告/錯誤跳機保護\n"
        "外型尺寸:1200*800*1280(mm)\n"
        "啟動電流衝擊小,恆壓恆溫控制,故障率低"
    ),
    "50馬力永磁變頻高效能補助專案型空壓機": (
        "型號:HCV-50PM-A\n馬力:5馬至50馬<隨壓力調整馬力>\n噪音:75±5\n"
        "電壓:三相220V\nIE4永磁高效率馬達<馬達無軸承><無皮帶>\n"
        "5kg~8kg可選變頻壓力\nLCD液晶顯示預警/警告/錯誤跳機保護\n"
        "外型尺寸:1300*1000*1450(mm)\n"
        "啟動電流衝擊小,恆壓恆溫控制,故障率低"
    ),
    "75馬力永磁變頻高效能補助專案型空壓機": (
        "型號:HCV-55PM-A\n馬力:10馬至75馬<隨壓力調整馬力>\n噪音:80±5\n"
        "電壓:三相220V\nIE4永磁高效率馬達<馬達無軸承><無皮帶>\n"
        "5kg~8kg可選變頻壓力\nLCD液晶顯示預警/警告/錯誤跳機保護\n"
        "外型尺寸:1900*1340*1700(mm)\n"
        "啟動電流衝擊小,恆壓恆溫控制,故障率低"
    ),
    "100馬力永磁變頻高效能補助專案型空壓機": (
        "型號:HCV-100PM-A\n馬力:15馬至100馬<隨壓力調整馬力>\n噪音:82±5\n"
        "電壓:三相220V\nIE4永磁高效率馬達\n"
        "5kg~8kg可選變頻壓力\nLCD液晶顯示預警/警告/錯誤跳機保護\n"
        "外型尺寸:2000*1340*1670(mm)\n"
        "啟動電流衝擊小,恆壓恆溫控制,故障率低"
    ),
}

# --- 3. 產品資料庫 (補齊乾燥機與配件) ---
products = {
    "空壓機": {
        "HCV高端系列": [
            ("10馬力永磁變頻高效能補助專案型空壓機", "air_10.png"), ("20馬力永磁變頻高效能補助專案型空壓機", "air_20.png"), 
            ("30馬力永磁變頻高效能補助專案型空壓機", "air_30.png"), ("50馬力永磁變頻高效能補助專案型空壓機", "air_50.png"),
            ("75馬力永磁變頻高效能補助專案型空壓機", "air_75.png"), ("100馬力永磁變頻高效能補助專案型空壓機", "air_100.png")
        ],
        "標準型": [("5馬空壓機", "air_5.png"),("10馬空壓機", "air_5.png"),("20馬空壓機", "air_5.png"),("30馬空壓機", "air_5.png"),
               ("50馬空壓機", "air_5.png"),("75馬空壓機", "air_5.png"),("100馬空壓機", "air_5.png"),]
    },
    "儲氣筒": [("155儲氣筒", "tank_105.png"), ("360儲氣筒", "tank_360.png"), ("660儲氣筒", "tank_660.png")],
    "乾燥機": {
        "宙升": [("5馬宙升乾燥機SD-005", "zs_dryer_5.png"), ("10馬宙升乾燥機SD-010", "zs_dryer_10.png"), ("15馬宙升乾燥機", "zs_dryer_15.png"), ("20馬宙升乾燥機", "zs_dryer_20.png"), ("30馬宙升乾燥機SD-030", "zs_dryer_30.png"), ("50馬宙升乾燥機SD-050", "zs_dryer_50.png"), ("100馬宙升乾燥機SD-100", "zs_dryer_100.png"),],
        "艾冷": [("5馬艾冷乾燥機", "al_dryer_5.png"), ("10馬艾冷乾燥機", "al_dryer_10.png"), ("20馬艾冷乾燥機", "al_dryer_20.png")]
    },
    "超精密過濾器組": {
        "合成牌": [("超精密過濾器組(合成牌)", "filter.png")],
        "PARK": [("超精密過濾器組(PARK)", "filter.png")]
    },
    "選配配件": [("Ckd自動排水器", "drainer_ckd.png"), ("電子式自動排水器", "drainer_e.png"),("過濾器濾心", "filter_core.png"),("旋風分離器", "separator.png")]
}

unit_map = {"155儲氣筒": "只", "360儲氣筒": "只", "660儲氣筒": "只", "Ckd自動排水器": "只","過濾器濾心": "只","旋風分離器": "只","電子式自動排水器": "只"}

if 'cart' not in st.session_state: st.session_state.cart = {}
if 'price_config' not in st.session_state:
    st.session_state.price_config = {n: 0 for series in products.values() for n, _ in (series if isinstance(series, list) else [i for sub in series.values() for i in sub])}

# --- 4. 產品選擇介面 ---
st.title("請選擇設備類別")
tabs = st.tabs(["空壓機", "儲氣筒", "乾燥機", "超精密過濾器組", "選配配件"])

def display_items(item_list):
    cols = st.columns(3)
    for i, (name, img) in enumerate(item_list):
        with cols[i % 3]:
            if os.path.exists(img): st.image(img, width=220)
            st.write(f"**{name}**")
            if st.button(f"➕ 加入報價單", key=f"btn_{name}"):
                st.session_state.cart[name] = st.session_state.cart.get(name, 0) + 1
                st.rerun()

with tabs[0]: display_items(products["空壓機"][st.radio("系列", ["HCV高端系列", "標準型"], horizontal=True)])
with tabs[1]: display_items(products["儲氣筒"])
with tabs[2]: display_items(products["乾燥機"][st.radio("品牌", ["宙升", "艾冷"], horizontal=True)])
with tabs[3]: display_items(products["超精密過濾器組"][st.radio("品牌 ", ["合成牌", "PARK"], horizontal=True)])
with tabs[4]: display_items(products["選配配件"])

# --- 5. 報價清單管理 ---
st.divider()
if st.session_state.cart:
    col_list, col_admin = st.columns([2, 1])
    with col_admin:
        with st.popover("⚙️ 修改客戶/價格", use_container_width=True):
            customer_name = st.text_input("客戶名稱")
            contact_person = st.text_input("聯絡人")
            voltage = st.radio("電力規格", ["220V", "380V"], horizontal=True)
            for name in sorted(st.session_state.cart.keys()):
                st.session_state.price_config[name] = st.number_input(f"{name} 單價", value=st.session_state.price_config.get(name, 0), step=100)

    with col_list:
        st.subheader("📋 目前報價清單")
        table_data = []
        total_val = 0
        for name, qty in st.session_state.cart.items():
            p = st.session_state.price_config.get(name, 0)
            sub = p * qty
            total_val += sub
            table_data.append([name, unit_map.get(name, "台"), qty, f"${p:,}", f"${sub:,}"])
        st.table(pd.DataFrame(table_data, columns=["品名及規格", "單位", "數量", "單價", "金額"]))

    # --- 6. Excel 匯出邏輯 ---
    template_path = "翌新估價單EXCELNEW.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        bold_kai = Font(name='標楷體', size=12, bold=True)
        spec_kai = Font(name='標楷體', size=11, bold=True)

        ws['B11'] = customer_name
        ws['B11'].font = bold_kai
        ws['B12'] = contact_person
        ws['B12'].font = bold_kai
        ws['H14'] = f"日期：{datetime.now().strftime('%Y-%m-%d')}"
        ws['H14'].font = bold_kai
        
        current_row = 17
        for i, (name, qty) in enumerate(st.session_state.cart.items()):
            p = st.session_state.price_config.get(name, 0)
            
            ws.cell(row=current_row, column=1, value=i+1).font = bold_kai
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
            name_cell = ws.cell(row=current_row, column=2, value=f"{name} ({voltage})")
            name_cell.font = bold_kai
            
            ws.cell(row=current_row, column=7, value=qty).font = bold_kai
            ws.cell(row=current_row, column=8, value=p).font = bold_kai
            ws.cell(row=current_row, column=9, value=p * qty).font = bold_kai
            
            if name in product_specs:
                current_row += 1
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
                spec_cell = ws.cell(row=current_row, column=2, value=product_specs[name])
                spec_cell.font = spec_kai
                spec_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                ws.row_dimensions[current_row].height = 200 # 維持 10 馬規格高度
            
            current_row += 1

        # --- I36 加粗 ---
        ws['I36'] = total_val
        ws['I36'].font = bold_kai 
        
        output = io.BytesIO()
        wb.save(output)
        
        c1, c2 = st.columns(2)
        with c1: st.download_button("📤 下載標楷體報價單", data=output.getvalue(), file_name=f"報價單_{customer_name}.xlsx", use_container_width=True)
        with c2:
            if st.button("🗑️ 清空重選", use_container_width=True):
                st.session_state.cart = {}
                st.rerun()
